import json
import os
from pathlib import Path
import openpyxl
from datetime import datetime

# Paths
ROOT = Path(__file__).resolve().parent
TEMPLATE_PATH = ROOT / "template.html"
OUT_PATH = ROOT / "index.html"

# Category rules (matching original script definitions for compatibility)
FUEL_KEYWORDS = ("주유소", "에너지", "오일뱅크", "석유", "주유")
TOLL_KEYWORDS = ("대교", "고속도로", "브릿지", "브리지", "도로")
TOLL_PAYMENT_METHOD = "현대카드ZERO Edition2 포인트형 하이패스"
CARD_ISSUER_BY_CONTENT = {
    "하나카드": "하나카드",
    "하나카드결제": "하나카드",
    "현대카드": "현대카드",
    "삼성카드": "삼성카드",
    "삼성카드(주)": "삼성카드",
}

def parse_amount(value):
    if value is None or value == "":
        return 0.0
    try:
        return float(str(value).replace(",", ""))
    except ValueError:
        return 0.0

def find_target_excel_file():
    target_dir = Path(r"C:\Users\home\Downloads\조병철님_2025-07-02~2026-07-02 (2)")
    if not target_dir.exists():
        return None
    
    excel_files = list(target_dir.glob("*.xlsx"))
    if not excel_files:
        return None
    
    excel_files.sort(key=lambda f: f.stat().st_mtime, reverse=True)
    return excel_files[0]

def main():
    excel_path = find_target_excel_file()
    transactions = []
    
    if not excel_path:
        print("Warning: Target excel file not found in Downloads folder. Falling back to default empty list.")
    else:
        print(f"Reading transactions from excel: {excel_path}")
        try:
            wb = openpyxl.load_workbook(excel_path, data_only=True)
            sheet_name = "가계부 내역"
            if sheet_name not in wb.sheetnames:
                sheet_name = wb.sheetnames[0]
            
            sheet = wb[sheet_name]
            rows = list(sheet.iter_rows(values_only=True))
            
            if len(rows) > 1:
                headers = [h.strip() if h else "" for h in rows[0]]
                col_idx = {h: i for i, h in enumerate(headers) if h}
                
                for r in rows[1:]:
                    if not r or all(cell is None for cell in r):
                        continue
                    
                    raw_date = r[col_idx.get("날짜")] if "날짜" in col_idx else ""
                    if isinstance(raw_date, datetime):
                        date_str = raw_date.strftime("%Y-%m-%d")
                    else:
                        date_str = str(raw_date) if raw_date else ""
                        
                    raw_time = r[col_idx.get("시간")] if "시간" in col_idx else ""
                    if isinstance(raw_time, datetime):
                        time_str = raw_time.strftime("%H:%M:%S")
                    elif hasattr(raw_time, "strftime"):
                        time_str = raw_time.strftime("%H:%M:%S")
                    else:
                        time_str = str(raw_time) if raw_time else ""
                        
                    raw_amount = parse_amount(r[col_idx.get("금액")] if "금액" in col_idx else 0)
                    
                    tx = {
                        "date": date_str,
                        "time": time_str,
                        "type": str(r[col_idx.get("타입")] if "타입" in col_idx else ""),
                        "category": str(r[col_idx.get("대분류")] if "대분류" in col_idx else "미분류"),
                        "subcategory": str(r[col_idx.get("소분류")] if "소분류" in col_idx else "미분류"),
                        "content": str(r[col_idx.get("내용")] if "내용" in col_idx else ""),
                        "amount": raw_amount,
                        "currency": str(r[col_idx.get("화폐")] if "화폐" in col_idx else "KRW"),
                        "payment": str(r[col_idx.get("결제수단")] if "결제수단" in col_idx else ""),
                        "memo": str(r[col_idx.get("메모")] if "메모" in col_idx else "")
                    }
                    transactions.append(tx)
        except Exception as e:
            print(f"Error parsing excel file: {e}")

    # Sort transactions
    transactions.sort(key=lambda x: (x.get("date", ""), x.get("time", "")), reverse=True)

    # Convert to JSON for injection
    tx_json = json.dumps(transactions, ensure_ascii=False)
    fuel_kw_json = json.dumps(FUEL_KEYWORDS, ensure_ascii=False)
    toll_kw_json = json.dumps(TOLL_KEYWORDS, ensure_ascii=False)
    card_iss_json = json.dumps(CARD_ISSUER_BY_CONTENT, ensure_ascii=False)

    # Read template HTML
    if not TEMPLATE_PATH.exists():
        print(f"Error: Template file not found at {TEMPLATE_PATH}")
        return

    html_content = TEMPLATE_PATH.read_text(encoding="utf-8")

    # Inject variables via string replacement
    supabase_url = os.environ.get("SUPABASE_URL", "__SUPABASE_URL__")
    supabase_anon_key = os.environ.get("SUPABASE_ANON_KEY", "__SUPABASE_ANON_KEY__")

    html_content = html_content.replace("__SUPABASE_URL__", supabase_url)
    html_content = html_content.replace("__SUPABASE_ANON_KEY__", supabase_anon_key)
    html_content = html_content.replace("__TRANSACTIONS_JSON__", tx_json)
    html_content = html_content.replace("__FUEL_KEYWORDS_JSON__", fuel_kw_json)
    html_content = html_content.replace("__TOLL_KEYWORDS_JSON__", toll_kw_json)
    html_content = html_content.replace("__TOLL_PAYMENT_METHOD__", TOLL_PAYMENT_METHOD)
    html_content = html_content.replace("__CARD_ISSUERS_JSON__", card_iss_json)

    # Write to HTML output
    OUT_PATH.write_text(html_content, encoding="utf-8")
    print(f"Success: Generated interactive dashboard at {OUT_PATH}")

if __name__ == "__main__":
    main()
